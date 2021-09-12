# 对excel进行操作
from API_practice.data_handle.copy_excel import copy_excel
import openpyxl


class OperateExcel:
    def __init__(self, file_path=None, sheet_id=None):
        # 生成一个新excel来记录测试结果
        self.new_excel = copy_excel()
        if file_path:
            self.file_path = file_path
            self.sheet_id = sheet_id
        else:
            self.file_path = self.new_excel
            self.sheet_id = 0
        self.sheet_contents = self.get_sheet()

    # 获取sheet内容
    def get_sheet(self):
        wb = openpyxl.load_workbook(self.file_path)
        table = wb.worksheets[self.sheet_id]
        return table

    # 获取单元格的列数
    def get_sheet_rows(self):
        return self.sheet_contents.max_row

    # 获取单元格的行数
    def get_sheet_column(self):
        return self.sheet_contents.max_column

    # 获取单元格内容
    def get_cell_value(self, row, col):
        return self.sheet_contents.cell(row, col).value

    def write_result(self, row, col, value):
        wb = openpyxl.load_workbook(self.file_path)
        wb1 = wb.active.cell(row, col, value)
        wb.save(self.file_path)

# if __name__ == '__main__':
#     open_excel = OperateExcel()
#     rows = open_excel.get_sheet_rows()
#     column = open_excel.get_sheet_column()
#     cell = open_excel.get_cell_content(2,3)
#     open_excel.write_result(17, 2, 'test')
#     print(rows, column, cell)
#     print(open_excel.get_sheet_rows())
